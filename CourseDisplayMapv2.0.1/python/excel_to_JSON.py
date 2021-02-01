# Written by Brenton Storm Feb-1-2021
from openpyxl import load_workbook
import json

#Loads in the Excel file
wb = load_workbook('../testFiles/Fall 2020 & Sprng 2021_FINAL.xlsx')

#creates a list of sheet names
sheetNames = []
for i in wb.sheetnames:
    sheetNames.append(i)

#creates a list of sheets by names
sheets = []
for i in sheetNames:
    sheets.append(wb[i])

#finds the max rows and columns
maxRow = []
maxCol = []
for i in sheets:
    maxRow.append(i.max_row)
    maxCol.append(i.max_column)

#sets up JSON
data = {}
for i in sheets:
    data[i.title] = []

#parses into JSON each sheet
for i in sheets:
    temp = data[i.title]
    #for each row
    for j in range(2, i.max_row+1):
        rows = {}
        #for each 'key: value' pair in a row
        for k in range(1, i.max_column+1):
            key = i.cell(row=1, column=k).value
            value = i.cell(row=j, column=k).value
            y = {key: value}
            rows.update(y)
        temp.append(rows)

#writes out to file
with open('../JSON.txt', 'w') as outfile:
    json.dump(data, outfile, indent=4)

print('Success!')
